from __future__ import annotations

import argparse
import math
import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from difflib import SequenceMatcher

from openpyxl import load_workbook


TABLE_COLUMNS: list[str] = [
    "Verdict",
    "Confidence",
    "SupplierTitle",
    "AmazonTitle",
    "Supplier EAN",
    "Amazon EAN",
    "ASIN",
    "SupplierPrice",
    "SellingPrice",
    "NetProfit",
    "ROI",
    "Sales",
    "Pack Verdict",
    "Adjusted Profit",
    "Key Match Evidence",
    "Filter Reason",
]


TITLE_STOPWORDS: set[str] = {
    "a",
    "an",
    "and",
    "are",
    "as",
    "at",
    "by",
    "for",
    "from",
    "in",
    "is",
    "it",
    "of",
    "on",
    "or",
    "the",
    "to",
    "with",
    "without",
}

GENERIC_EVIDENCE_TOKENS: set[str] = {
    "pack",
    "pk",
    "set",
    "sets",
    "piece",
    "pieces",
    "pcs",
    "pair",
    "pairs",
    "assorted",
    "variety",
    "mixed",
    "refill",
    "replacement",
    "compatible",
    "size",
    "small",
    "medium",
    "large",
    "extra",
    "new",
}


LUXURY_IP_BRANDS: tuple[str, ...] = (
    "jo malone",
    "chanel",
    "dior",
    "gucci",
    "louis vuitton",
    "prada",
    "hermès",
    "hermes",
    "apple",
    "samsung",
    "sony",
    "microsoft",
    "nike",
    "adidas",
)


MEASUREMENT_UNITS: tuple[str, ...] = (
    "mm",
    "cm",
    "m",
    "in",
    "inch",
    "inches",
    "ft",
    "feet",
    "ml",
    "l",
    "litre",
    "liter",
    "g",
    "kg",
    "oz",
)

CATEGORY_KEYWORDS: dict[str, set[str]] = {
    "electronics": {
        "smartphone",
        "phone",
        "tablet",
        "laptop",
        "camera",
        "charger",
        "usb",
        "bluetooth",
        "android",
        "iphone",
        "motorola",
        "samsung",
        "gb",
        "ram",
    },
    "party": {
        "birthday",
        "balloon",
        "badge",
        "banner",
        "valentine",
        "christmas",
        "wrapping",
        "wrap",
        "candle",
        "candles",
        "decorations",
        "decoration",
        "cupcake",
    },
    "toys": {
        "toy",
        "toys",
        "blocks",
        "building",
        "model",
        "kit",
        "stem",
        "mould",
        "king",
        "engine",
    },
    "pet": {
        "dog",
        "cat",
        "pet",
        "petsafe",
        "launcher",
        "tennis",
        "litter",
        "collar",
        "leash",
    },
    "tools": {
        "tool",
        "tools",
        "drill",
        "screwdriver",
        "wrench",
        "pliers",
        "hammer",
        "saw",
        "socket",
        "tape",
        "measure",
    },
    "kitchen": {
        "bowl",
        "plate",
        "mug",
        "cup",
        "knife",
        "fork",
        "spoon",
        "pan",
        "kettle",
    },
}

DISTINCTIVE_KEYWORDS: set[str] = {
    "smartphone",
    "motorola",
    "iphone",
    "samsung",
    "petsafe",
    "balloon",
    "badge",
    "cupcake",
}


def dubai_today() -> str:
    dt = datetime.now(timezone.utc) + timedelta(hours=4)
    return dt.date().isoformat()


def to_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_to_digits(value: Any) -> str:
    s = to_str(value)
    if not s:
        return ""
    if "e" in s.lower():
        return ""
    s = s.replace(".0", "")
    return re.sub(r"\D", "", s)


def gtin_checksum_ok(digits: str) -> bool:
    if not digits.isdigit():
        return False
    n = len(digits)
    if n not in (8, 12, 13, 14):
        return False

    body = digits[:-1]
    check = int(digits[-1])

    total = 0
    for i, ch in enumerate(body[::-1], start=1):
        d = int(ch)
        total += d * (3 if i % 2 == 1 else 1)
    calc = (10 - (total % 10)) % 10
    return calc == check


def normalize_gtin(digits: str) -> str:
    if not digits.isdigit():
        return digits
    if len(digits) in (8, 12, 13, 14) and gtin_checksum_ok(digits):
        return digits
    for target_len in (12, 13, 14):
        if len(digits) < target_len:
            padded = digits.zfill(target_len)
            if gtin_checksum_ok(padded):
                return padded
    return digits


def is_strict_valid_barcode(digits: str) -> bool:
    if not isinstance(digits, str):
        return False
    if not digits.isdigit():
        return False
    normalized = normalize_gtin(digits)
    if len(normalized) not in (8, 12, 13, 14):
        return False
    if re.search(r"0{6,}$", normalized):
        return False
    return gtin_checksum_ok(normalized)


def title_similarity(supplier_title: str, amazon_title: str) -> float:
    if not supplier_title or not amazon_title:
        return 0.0
    return SequenceMatcher(None, supplier_title.lower(), amazon_title.lower()).ratio()


def tokenize_title(title: str) -> list[str]:
    if not title:
        return []
    tokens = re.findall(r"[a-z0-9]+", title.lower())
    return [
        t
        for t in tokens
        if len(t) >= 3 and t not in TITLE_STOPWORDS and t not in GENERIC_EVIDENCE_TOKENS
    ]


def shared_title_anchors(supplier_title: str, amazon_title: str, max_items: int = 6) -> list[str]:
    sup = tokenize_title(supplier_title)
    amz = tokenize_title(amazon_title)
    if not sup or not amz:
        return []
    shared = sorted(set(sup).intersection(amz))
    return shared[:max_items]

def classify_title_category(title: str) -> tuple[str | None, int]:
    tokens = set(tokenize_title(title))
    if not tokens:
        return None, 0
    best_cat: str | None = None
    best_score = 0
    for cat, kws in CATEGORY_KEYWORDS.items():
        score = sum(1 for k in kws if k in tokens)
        if score > best_score:
            best_cat = cat
            best_score = score
    if best_score >= 2:
        return best_cat, best_score
    if best_score == 1 and best_cat:
        if any(k in tokens for k in CATEGORY_KEYWORDS[best_cat].intersection(DISTINCTIVE_KEYWORDS)):
            return best_cat, best_score
    return None, 0


def parse_sales(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)) and not (isinstance(value, float) and math.isnan(value)):
        return int(value)
    s = to_str(value)
    m = re.search(r"\d+", s)
    return int(m.group(0)) if m else 0


def to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)) and not (isinstance(value, float) and math.isnan(value)):
        return float(value)
    s = to_str(value)
    s = s.replace("£", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def fmt_gbp(value: float) -> str:
    return f"£{value:.2f}"


def fmt_roi(value: float) -> str:
    return f"{value:.1f}%"


def fmt_int(value: int) -> str:
    return str(int(value))


def has_luxury_ip_brand(title: str) -> str | None:
    t = title.lower() if title else ""
    for b in LUXURY_IP_BRANDS:
        if b in t:
            return b
    return None


@dataclass(frozen=True)
class PackInfo:
    count: int | None
    evidence: str | None
    dimension_shield_triggered: bool


def _dimension_pattern() -> re.Pattern[str]:
    units = "|".join(re.escape(u) for u in MEASUREMENT_UNITS)
    return re.compile(
        rf"(\d+)\s*[x×]\s*(\d+)\s*(?:{units})\b",
        flags=re.IGNORECASE,
    )


def extract_pack_info(title: str) -> PackInfo:
    t = title.lower() if title else ""
    if not t:
        return PackInfo(count=None, evidence=None, dimension_shield_triggered=False)

    dim_pat = _dimension_pattern()
    if dim_pat.search(t):
        return PackInfo(count=None, evidence=None, dimension_shield_triggered=True)

    units = "|".join(re.escape(u) for u in MEASUREMENT_UNITS)
    # e.g., "3 x 200ml" => pack count is 3 (NOT 600ml)
    m = re.search(rf"\b(\d+)\s*[x×]\s*(\d+)\s*(?:{units})\b", t)
    if m:
        return PackInfo(count=int(m.group(1)), evidence=m.group(0), dimension_shield_triggered=False)

    # e.g., "3 x 15 pack" => 45
    m = re.search(r"\b(\d+)\s*[x×]\s*(\d+)\s*(pack|pk|pcs|pieces|pairs|rolls|bags|cases?)\b", t)
    if m:
        a = int(m.group(1))
        b = int(m.group(2))
        if 1 < a <= 500 and 1 < b <= 500 and a * b <= 500:
            return PackInfo(count=a * b, evidence=m.group(0), dimension_shield_triggered=False)

    patterns: list[re.Pattern[str]] = [
        re.compile(r"\bpack of (\d+)\b", flags=re.IGNORECASE),
        re.compile(r"\bset of (\d+)\b", flags=re.IGNORECASE),
        re.compile(
            r"\b(\d+)\s*(pack|pk|pcs|pieces|pairs|rolls|bags|cases?|pce|pces|ct|count)\b",
            flags=re.IGNORECASE,
        ),
        re.compile(r"\bpk\s*(\d+)\b", flags=re.IGNORECASE),
        re.compile(r"\bpk(\d+)\b", flags=re.IGNORECASE),
        re.compile(r"\b(\d+)\s*pce\b", flags=re.IGNORECASE),
        re.compile(r"\((\d+)\s*(pack|pk)\)", flags=re.IGNORECASE),
        # leading "3 x ..." without units; treat as possibly multipack (uncertain but useful)
        re.compile(r"^\s*(\d+)\s*[x×]\s*", flags=re.IGNORECASE),
    ]

    for pat in patterns:
        m = pat.search(t)
        if not m:
            continue
        qty = int(m.group(1))
        if 1 < qty <= 500:
            return PackInfo(count=qty, evidence=m.group(0).strip(), dimension_shield_triggered=False)

    return PackInfo(count=None, evidence=None, dimension_shield_triggered=False)


def truncate(text: str, max_len: int) -> str:
    if len(text) <= max_len:
        return text
    if max_len <= 3:
        return text[:max_len]
    return text[: max_len - 3] + "..."


def format_fixed_width_table(rows: list[dict[str, str]]) -> str:
    caps: dict[str, int] = {
        # Do not truncate verdict labels (e.g., "NEEDS VERIFICATION").
        "Verdict": 25,
        "Confidence": 10,
        "SupplierTitle": 110,
        "AmazonTitle": 120,
        "Supplier EAN": 14,
        "Amazon EAN": 14,
        "ASIN": 10,
        "SupplierPrice": 13,
        "SellingPrice": 13,
        "NetProfit": 11,
        "ROI": 10,
        "Sales": 7,
        "Pack Verdict": 55,
        "Adjusted Profit": 15,
        "Key Match Evidence": 70,
        "Filter Reason": 90,
    }

    def cell(col: str, val: str) -> str:
        cap = caps.get(col)
        if cap is None:
            return val
        return truncate(val, cap)

    # Clamp cells first so width calc respects caps.
    clamped_rows: list[dict[str, str]] = []
    for r in rows:
        clamped_rows.append({c: cell(c, r.get(c, "")) for c in TABLE_COLUMNS})

    widths: dict[str, int] = {}
    for c in TABLE_COLUMNS:
        widths[c] = max(len(c), *(len(r.get(c, "")) for r in clamped_rows))

    header = "| " + " | ".join(c.ljust(widths[c]) for c in TABLE_COLUMNS) + " |"
    sep = "|-" + "-|-".join("-" * widths[c] for c in TABLE_COLUMNS) + "-|"
    lines = [header, sep]
    for r in clamped_rows:
        lines.append("| " + " | ".join(r.get(c, "").ljust(widths[c]) for c in TABLE_COLUMNS) + " |")
    return "\n".join(lines)


def infer_supplier_from_urls(urls: list[str]) -> str:
    hosts: list[str] = []
    for u in urls:
        if not u:
            continue
        try:
            host = urlparse(u).netloc.lower()
        except Exception:
            continue
        if host:
            hosts.append(host)
    if not hosts:
        return "-"
    return Counter(hosts).most_common(1)[0][0]


def load_sheet_as_rows(input_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(input_path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    if not headers or any(h is None for h in headers):
        raise ValueError("Missing headers in row 1")
    col_count = len(headers)

    rows: list[dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        item = dict(zip(headers, row[:col_count], strict=False))
        item["RowID"] = i
        rows.append(item)
    return rows


def build_report_rows(raw_rows: list[dict[str, Any]]) -> tuple[list[dict[str, str]], dict[str, int], str]:
    supplier = infer_supplier_from_urls([to_str(r.get("SupplierURL")) for r in raw_rows])
    verdict_counts: dict[str, int] = {"VERIFIED": 0, "HIGHLY LIKELY": 0, "NEEDS VERIFICATION": 0, "FILTERED OUT": 0}
    out_rows: list[dict[str, str]] = []

    for r in raw_rows:
        row_id = int(r.get("RowID", 0) or 0)
        supplier_title = to_str(r.get("SupplierTitle"))
        amazon_title = to_str(r.get("AmazonTitle"))
        asin = to_str(r.get("ASIN"))

        sales = parse_sales(r.get("bought_in_past_month"))
        supplier_price = to_float(r.get("SupplierPrice_incVAT"))
        selling_price = to_float(r.get("SellingPrice_incVAT"))
        net_profit = to_float(r.get("NetProfit"))
        roi = to_float(r.get("ROI"))

        sup_digits = clean_to_digits(r.get("EAN"))
        amz_digits = clean_to_digits(r.get("EAN_OnPage"))
        sup_norm = normalize_gtin(sup_digits) if sup_digits else ""
        amz_norm = normalize_gtin(amz_digits) if amz_digits else ""
        sup_valid = is_strict_valid_barcode(sup_norm) if sup_norm else False
        amz_valid = is_strict_valid_barcode(amz_norm) if amz_norm else False
        is_exact_ean_strict = sup_valid and amz_valid and sup_norm == amz_norm

        supplier_ean_disp = sup_norm if sup_valid else "-"
        amazon_ean_disp = amz_norm if amz_valid else "-"

        sim = title_similarity(supplier_title, amazon_title)
        anchors = shared_title_anchors(supplier_title, amazon_title)
        sup_cat, sup_cat_score = classify_title_category(supplier_title)
        amz_cat, amz_cat_score = classify_title_category(amazon_title)

        sup_pack = extract_pack_info(supplier_title)
        amz_pack = extract_pack_info(amazon_title)

        pack_verdict = "Pack unclear - verify manually"
        adjusted_profit = net_profit
        pack_reason: str | None = None

        pack_from_one_side = bool(sup_pack.count) ^ bool(amz_pack.count)
        no_pack_info = (
            not sup_pack.count
            and not amz_pack.count
            and not (sup_pack.dimension_shield_triggered or amz_pack.dimension_shield_triggered)
        )

        if sup_pack.dimension_shield_triggered or amz_pack.dimension_shield_triggered:
            pack_verdict = "1:1 Match (Dimension Shield)"
        elif sup_pack.count and amz_pack.count:
            if sup_pack.count == amz_pack.count:
                pack_verdict = "1:1 Match"
            else:
                ratio = amz_pack.count / sup_pack.count
                if ratio > 1.0:
                    near_int = abs(ratio - round(ratio)) <= 0.05
                    if near_int:
                        rsu = int(round(ratio))
                        pack_verdict = f"BUNDLE (Amazon {amz_pack.count} vs Supplier {sup_pack.count}); RSU={rsu}"
                        adjusted_profit = net_profit - (supplier_price * (rsu - 1))
                        pack_reason = f"Pack mismatch: need {rsu} supplier units for 1 Amazon unit"
                    else:
                        pack_verdict = f"Pack mismatch ambiguous (Amazon {amz_pack.count} vs Supplier {sup_pack.count})"
                        pack_reason = "Pack math uncertain - verify manually"
                else:
                    pack_verdict = f"SPLIT risk (Supplier {sup_pack.count} vs Amazon {amz_pack.count})"
                    pack_reason = "Supplier seems to be multipack; Amazon seems single - verify repack compliance"
        elif sup_pack.count or amz_pack.count:
            if sup_pack.count:
                pack_verdict = f"Supplier pack={sup_pack.count}; Amazon pack unclear"
            else:
                pack_verdict = f"Amazon pack={amz_pack.count}; Supplier pack unclear"
            pack_reason = "Pack math uncertain - verify manually"
        elif is_exact_ean_strict and no_pack_info:
            # For strict exact EAN rows with no pack indicators, default to 1:1 match.
            pack_verdict = "1:1 Match (Exact EAN default)"

        # Evidence must be row-grounded: tokens from titles OR strict exact EAN.
        if is_exact_ean_strict:
            evidence = "Exact EAN match"
            if anchors:
                evidence += "; shared: " + ", ".join(anchors[:4])
        else:
            evidence = "shared: " + ", ".join(anchors[:6]) if anchors else "-"

        filter_reason = "-"
        confidence: int
        pack_uncertain = (
            pack_reason is not None
            or pack_from_one_side
            or any(s in pack_verdict.lower() for s in ("split risk", "ambiguous"))
        )

        if is_exact_ean_strict:
            confidence = 95
            verdict = "VERIFIED"
            if sim < 0.20 and not anchors:
                confidence = 85
                filter_reason = "Title mismatch despite exact EAN - verify extraction integrity"
                verdict = "NEEDS VERIFICATION"
            if sup_cat and amz_cat and sup_cat != amz_cat:
                # Contradiction: don't discard the match outright, but flag for manual review.
                confidence = min(confidence, 90)
                filter_reason = f"Category mismatch despite exact EAN ({sup_cat} vs {amz_cat}) - verify variant/listing"
                verdict = "NEEDS VERIFICATION"
            if verdict == "VERIFIED" and pack_uncertain:
                verdict = "NEEDS VERIFICATION"
                filter_reason = pack_reason or "Pack math uncertain - verify manually"
            if sales <= 0:
                verdict = "NEEDS VERIFICATION"
                filter_reason = "Sales=0 (or missing); verify demand/Keepa"
            elif net_profit <= 0:
                verdict = "FILTERED OUT"
                filter_reason = "NetProfit<=0 (not profitable)"
            elif adjusted_profit <= 0:
                verdict = "FILTERED OUT"
                filter_reason = "Profit-after-pack-sanity<=0 (not profitable after RSU)"
            elif verdict == "VERIFIED":
                filter_reason = "-"
        else:
            # Recall-first: prefer NEEDS VERIFICATION unless contradiction is clear.
            category_mismatch = sup_cat and amz_cat and sup_cat != amz_cat and (sup_cat_score + amz_cat_score) >= 2
            if category_mismatch:
                verdict = "FILTERED OUT"
                confidence = 10
                filter_reason = f"Category mismatch trap ({sup_cat} vs {amz_cat})"
            elif sim < 0.16 and not anchors:
                verdict = "FILTERED OUT"
                confidence = 5
                filter_reason = "Clear mismatch: no shared anchors + very low title similarity"
            else:
                base = int(round(sim * 100))
                confidence = max(10, min(90, base))
                if len(anchors) >= 2:
                    confidence = min(92, confidence + 8)

                strong_match = (sim >= 0.55 and len(anchors) >= 2) or (sim >= 0.70)
                if strong_match:
                    verdict = "HIGHLY LIKELY"
                else:
                    verdict = "NEEDS VERIFICATION"

                ip_brand = has_luxury_ip_brand(supplier_title) or has_luxury_ip_brand(amazon_title)
                if ip_brand:
                    verdict = "NEEDS VERIFICATION"
                    filter_reason = f"Potential IP risk brand detected ({ip_brand}); verify gating"

                if verdict == "HIGHLY LIKELY":
                    if pack_reason:
                        verdict = "NEEDS VERIFICATION"
                        filter_reason = pack_reason
                    elif sales <= 0:
                        verdict = "NEEDS VERIFICATION"
                        filter_reason = "Sales=0 (or missing); verify demand/Keepa"
                    elif net_profit <= 0:
                        verdict = "FILTERED OUT"
                        filter_reason = "NetProfit<=0 (not profitable)"
                    elif adjusted_profit <= 0:
                        verdict = "FILTERED OUT"
                        filter_reason = "Profit-after-pack-sanity<=0 (not profitable after RSU)"
                    else:
                        filter_reason = "-"
                elif verdict == "NEEDS VERIFICATION" and filter_reason == "-":
                    if pack_reason:
                        filter_reason = pack_reason
                    elif sim < 0.30 and len(anchors) <= 1:
                        filter_reason = "Weak title anchors; verify variant/brand/product type"
                    else:
                        filter_reason = "Verify pack/variant/EAN on page; match looks plausible"

        verdict_counts[verdict] += 1

        out_rows.append(
            {
                "Verdict": verdict,
                "Confidence": fmt_int(confidence),
                "SupplierTitle": f"[{row_id}] {supplier_title}",
                "AmazonTitle": amazon_title,
                "Supplier EAN": supplier_ean_disp,
                "Amazon EAN": amazon_ean_disp,
                "ASIN": asin,
                "SupplierPrice": fmt_gbp(supplier_price),
                "SellingPrice": fmt_gbp(selling_price),
                "NetProfit": fmt_gbp(net_profit),
                "ROI": fmt_roi(roi),
                "Sales": fmt_int(sales),
                "Pack Verdict": pack_verdict,
                "Adjusted Profit": fmt_gbp(adjusted_profit),
                "Key Match Evidence": evidence,
                "Filter Reason": filter_reason,
            }
        )

    return out_rows, verdict_counts, supplier


def write_report(
    output_path: Path,
    input_path: Path,
    supplier: str,
    rows: list[dict[str, str]],
    verdict_counts: dict[str, int],
) -> None:
    def section(name: str) -> str:
        return f"{name} (count={verdict_counts.get(name, 0)})"

    by_verdict: dict[str, list[dict[str, str]]] = {k: [] for k in verdict_counts}
    for r in rows:
        by_verdict[r["Verdict"]].append(r)

    by_verdict["VERIFIED"].sort(key=lambda x: (int(x["Sales"]), float(x["Adjusted Profit"].replace("£", ""))), reverse=True)
    by_verdict["HIGHLY LIKELY"].sort(key=lambda x: (int(x["Confidence"]), int(x["Sales"])), reverse=True)
    by_verdict["NEEDS VERIFICATION"].sort(key=lambda x: (int(x["Confidence"]), int(x["Sales"])), reverse=True)

    total_analyzed = len(rows)

    narrative = (
        "This report is recall-first: anything not explicitly contradicted is routed to NEEDS VERIFICATION.\n"
        "Long titles are truncated inside tables for readability; the RowID prefix preserves traceability.\n"
        "VERIFIED and HIGHLY LIKELY sections enforce Sales>0 and positive profitability; all other rows are "
        "retained for audit in NEEDS VERIFICATION or FILTERED OUT.\n"
    )

    content: list[str] = []
    content.append("# PHASEA MANUAL REPORT")
    content.append("")
    content.append(f"**Generated:** {dubai_today()}")
    content.append(f"**Input File:** {str(input_path)}")
    content.append(f"**Supplier:** {supplier}")
    content.append("")
    content.append("## Summary Counts")
    content.append(f"- VERIFIED: {verdict_counts.get('VERIFIED', 0)}")
    content.append(f"- HIGHLY LIKELY: {verdict_counts.get('HIGHLY LIKELY', 0)}")
    content.append(f"- NEEDS VERIFICATION: {verdict_counts.get('NEEDS VERIFICATION', 0)}")
    content.append(f"- FILTERED OUT: {verdict_counts.get('FILTERED OUT', 0)}")
    content.append(f"- TOTAL ANALYZED: {total_analyzed}")
    content.append("")
    content.append(narrative.strip())
    content.append("")

    for v in ("VERIFIED", "HIGHLY LIKELY", "NEEDS VERIFICATION", "FILTERED OUT"):
        content.append(f"## {section(v)}")
        content.append("")
        content.append("```text")
        content.append(format_fixed_width_table(by_verdict[v]))
        content.append("```")
        content.append("")

    content.append("## Reconciliation")
    content.append(f"- Total rows in input: {total_analyzed}")
    content.append(
        "- Rows analyzed and categorized: "
        f"{verdict_counts.get('VERIFIED', 0)} + {verdict_counts.get('HIGHLY LIKELY', 0)} + "
        f"{verdict_counts.get('NEEDS VERIFICATION', 0)} + {verdict_counts.get('FILTERED OUT', 0)}"
    )
    content.append("")

    output_path.write_text("\n".join(content), encoding="utf-8")


def validate_output(rows: list[dict[str, str]]) -> list[str]:
    issues: list[str] = []
    for r in rows:
        verdict = r["Verdict"]
        sales = int(r["Sales"])
        net_profit = float(r["NetProfit"].replace("£", ""))
        adj_profit = float(r["Adjusted Profit"].replace("£", ""))
        if verdict in ("VERIFIED", "HIGHLY LIKELY"):
            if sales <= 0:
                issues.append(f"{verdict} row has Sales<=0: ASIN={r['ASIN']}")
            if net_profit <= 0:
                issues.append(f"{verdict} row has NetProfit<=0: ASIN={r['ASIN']}")
            if adj_profit <= 0:
                issues.append(f"{verdict} row has Adjusted Profit<=0: ASIN={r['ASIN']}")
        if verdict == "VERIFIED":
            if r["Supplier EAN"] == "-" or r["Amazon EAN"] == "-" or r["Supplier EAN"] != r["Amazon EAN"]:
                issues.append(f"VERIFIED row is not strict exact EAN: ASIN={r['ASIN']}")
    return issues


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    input_path: Path = args.input
    output_path: Path = args.output

    if not input_path.exists():
        raise SystemExit(f"Missing input: {input_path}")

    raw_rows = load_sheet_as_rows(input_path)
    report_rows, verdict_counts, supplier = build_report_rows(raw_rows)

    issues = validate_output(report_rows)
    if issues:
        # Keep going to allow inspection, but make it visible in stderr.
        print("VALIDATION_ISSUES:", len(issues))
        for i in issues[:50]:
            print(" -", i)

    write_report(output_path, input_path, supplier, report_rows, verdict_counts)
    print(f"WROTE {output_path} ({output_path.stat().st_size} bytes)")
    print("COUNTS", verdict_counts)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
